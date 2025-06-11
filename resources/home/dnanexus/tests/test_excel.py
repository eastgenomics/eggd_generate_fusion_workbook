"""
Tests for excel.py module
"""

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from utils import excel


class TestExcelUtils:
    """Unit tests for Excel utility functions in utils.excel."""

    @pytest.fixture
    def worksheet(self):
        """Fixture providing a sample worksheet with test data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        data = [
            ["Header1", "Header2", "Header3"],
            [1, 2, 3],
            [4, 5, 6],
            [7, 8, 9],
        ]
        for row in data:
            ws.append(row)
        return ws

    def test_add_extra_columns(self, worksheet):
        """Test inserting an extra column with a formula."""
        excel.add_extra_columns(worksheet, {"ExtraCol": "=A{row}+1"})
        assert worksheet.cell(1, 1).value == "ExtraCol"
        assert worksheet.cell(2, 1).value == "=A2+1"

    def test_apply_header_format(self, worksheet):
        """Test applying bold formatting to header row."""
        excel.apply_header_format(worksheet)
        assert worksheet["A1"].font.bold is True
        assert worksheet["B1"].font.bold is True

    def test_adjust_column_widths(self, worksheet):
        """Test adjusting column widths based on content length."""
        worksheet["A2"] = "short"
        worksheet["B2"] = "a much longer string"
        excel.adjust_column_widths(worksheet)
        # Will be improve to get the actual width in excel units
        assert (
            worksheet.column_dimensions["B"].width
            > worksheet.column_dimensions["A"].width
        )

    def test_set_tab_color(self, worksheet):
        """Test setting worksheet tab color."""
        excel.set_tab_color(worksheet, "FFFF00FF")
        assert worksheet.sheet_properties.tabColor.rgb == "FFFF00FF"

    def test_style_borders(self, worksheet):
        """Test applying borders to cells."""
        excel.style_borders(worksheet)
        assert worksheet["A2"].border.left.style == "thin"
        assert worksheet["B2"].border.bottom.style == "thin"

    def test_align_column_cells(self, worksheet):
        """Test aligning all cells in a column."""
        excel.align_column_cells(worksheet, "A", "center")
        for cell in worksheet["A"]:
            assert cell.alignment.horizontal == "center"

    def test_set_column_width(self, worksheet):
        """Test setting explicit column width."""
        excel.set_column_width(worksheet, "B", 25.0)
        assert worksheet.column_dimensions["B"].width == 25.0

    def test_get_col_letter(self, worksheet):
        """Test retrieving Excel column letter from a header name."""
        assert excel.get_col_letter(worksheet, "Header2") == "B"

    def test_drop_column(self, worksheet):
        """Test removing a column by header name."""
        result = excel.drop_column(worksheet, "Header2")
        assert result is True
        assert worksheet.cell(1, 2).value != "Header2"

    def test_rotate_headers(self, worksheet):
        """Test rotating header text."""
        excel.rotate_headers(worksheet, degree=45)
        assert worksheet["A1"].alignment.textRotation == 45

    def test_format_columns_to_two_dp(self, worksheet):
        """Test applying two-decimal format to numeric columns."""
        worksheet["A1"].value = "Days"
        worksheet["A2"].value = 12.3456
        excel.format_columns_to_two_dp(worksheet)
        assert worksheet["A2"].number_format == "0.00"

    def test_add_hyperlink(self, worksheet):
        """Test adding hyperlink formula to a cell."""
        cell = worksheet["A2"]
        excel.add_hyperlink(cell, "http://example.com", "Click")
        assert cell.value == '=HYPERLINK("http://example.com", "Click")'

    def test_add_drop_down_col(self, worksheet):
        """Test inserting dropdown data validation column."""
        excel.add_drop_down_col(
            worksheet,
            header_name="Review",
            dropdown_options=["Pass", "Fail"],
            prompt="Select result",
            title="Review",
            start=1,
            position=2,
        )
        assert worksheet["B1"].value == "Review"
        assert worksheet.column_dimensions["B"].width >= len("Review") + 2

    def test_colour_hyperlinks(self, worksheet):
        """Test formatting font color of hyperlink cells."""
        worksheet["A2"].value = '=HYPERLINK("http://example.com", "Example")'
        excel.colour_hyperlinks(worksheet)
        assert worksheet["A2"].font.color.rgb.lower().endswith("00007f")
