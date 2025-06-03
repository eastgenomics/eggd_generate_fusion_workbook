"""Utilities for formating summary sheet
"""

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .excel import (
    alternate_specimen_colors,
    highlight_specimen_borders,
    style_borders,
    set_column_width,
    align_column_cells,
    add_drop_down_col,
    get_col_letter,
    write_df_to_sheet,
    drop_column,
    rotate_headers,
)
from .utils import validate_config


def add_databar_to_ffpm(
    worksheet: Worksheet, df: pd.DataFrame, ffpm_col: str, index_col: int
) -> None:
    """
    Conditionally add databar to FFPM col (groupby specimen).

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where conditional formatting will be applied
    df : pandas.DataFrame
       Data frame containing original summary data
    ffpm_col : str
        Name of the FFPM column in the DataFrame
    index_col : str
        Name of the column to use for grouping ffpm

    Returns
    -------
    None
    """
    # Get col letter for FFPM
    col_letter = get_col_letter(worksheet, ffpm_col)

    for specimen, group in df.groupby(index_col):
        # adjust for excel index (1-based, +1 for headers)
        start = group.index[0] + 2
        end = group.index[-1] + 2

        cell_range = f"{col_letter}{start}:{col_letter}{end}"
        max_ffpm = group[ffpm_col].max()

        rule = openpyxl.formatting.rule.DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=max_ffpm,
            color="FFC854",
            showValue=True,
        )
        worksheet.conditional_formatting.add(cell_range, rule)

    # adjust cells to display databar correctly
    set_column_width(worksheet, col_letter, 10.0)
    align_column_cells(worksheet, col_letter)


def add_lookup_columns(
    worksheet: Worksheet, cols: dict, specimen_col: str = "B"
) -> None:
    """
    Adds formula columns while preserving merged specimen structure

    Parameters
    ----------
    worksheet : Worksheet
        Target Excel worksheet
    cols : dict
        {column_header: formula_template} mapping
    specimen_col : str
        Column letter containing merged specimen cells (default 'B')
    """
    # Get specimen column index
    spec_col_idx = worksheet[specimen_col][0].column

    # Insert new columns after specimen column
    worksheet.insert_cols(spec_col_idx + 1, amount=len(cols))

    # Get all merged ranges in specimen column
    merged_ranges = [
        r for r in worksheet.merged_cells.ranges if r.min_col == spec_col_idx
    ]

    # Keep track of rows in merged ranges
    merged_rows = set()
    for r in merged_ranges:
        merged_rows.update(range(r.min_row, r.max_row + 1))

    # Process each new column
    for col_offset, (header, formula_template) in enumerate(cols.items(), start=1):
        new_col_idx = spec_col_idx + col_offset
        new_col = openpyxl.utils.get_column_letter(new_col_idx)

        worksheet.cell(row=1, column=new_col_idx, value=header)

        # Apply formulas to merged cell groups
        for merged_range in merged_ranges:

            start = merged_range.min_row
            end = merged_range.max_row

            # Create merged area in new column
            worksheet.merge_cells(f"{new_col}{start}:{new_col}{end}")

            # Apply formula to first cell
            formula = formula_template.replace("{row}", str(start))
            new_cell = worksheet.cell(row=start, column=new_col_idx, value=formula)
            new_cell.alignment = Alignment(vertical="top")

            # Clear other merged cells (they inherit from first cell)
            for row in range(start + 1, end + 1):
                worksheet.cell(row=row, column=new_col_idx, value=None)

        # Handle rows (not part of any merged range). Case of 1 fusion per sample
        for row in range(2, worksheet.max_row + 1):
            if row not in merged_rows:
                formula = formula_template.replace("{row}", str(row))
                cell = worksheet.cell(row=row, column=new_col_idx, value=formula)


def format_summary_sheet(
    worksheet: Worksheet,
    source_df: pd.DataFrame,
    col_widths: dict,
    ffpm_col: str = "FFPM",
    index_col: str = "SPECIMEN",
) -> None:
    """
    Conditionally formats summary sheet rows per specimen.

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where conditional formatting will be applied
    source_df : pd.DataFrame
        The source data written to sheet
    col_widths : dict
        Mapping of col letter to a fix width size
    ffpm_col : str, optional
        Name of the FFPM column in the DataFrame, default is "FFPM"
    index_col : str, optional
        Name of the column to use for grouping ffpm, default is "SPECIMEN"

    Returns
    -------
    None
    """

    df = source_df.copy()
    if isinstance(df.index, pd.MultiIndex):
        df.reset_index(inplace=True)

    style_borders(worksheet)
    add_databar_to_ffpm(worksheet, df, ffpm_col, index_col)
    alternate_specimen_colors(worksheet, df, index_col)
    highlight_specimen_borders(worksheet, df, index_col)
    rotate_headers(worksheet)

    for col, width in col_widths.items():
        set_column_width(worksheet, col, width)


def write_summary(
    writer: pd.ExcelWriter,
    data: pd.DataFrame,
    config: dict,
) -> None:
    """Writes summary data to summary sheet with specified formating
    Parameters
    ----------
    writer : pd.ExcelWriter
        The Excel writer object to write the sheet into
    data : pd.DataFrame
        Source data containing pivot summary
    config : dict
        Config specifing styling and formating of sheet
    """
    expected_keys = [
        "sheet_name",
        "tab_color",
        "extra_cols",
        "drop_downs",
        "col_widths",
    ]
    validate_config(config, expected_keys)

    sheet_name = config["sheet_name"]
    tab_color = config["tab_color"]
    lookup_cols = config["extra_cols"]
    col_widths = config["col_widths"]

    write_df_to_sheet(
        writer,
        data,
        sheet_name=sheet_name,
        tab_color=tab_color,
        include_index=True,
    )
    summary_sheet = writer.sheets[sheet_name]
    drop_column(summary_sheet, "LEFTRIGHT")

    add_lookup_columns(summary_sheet, lookup_cols)

    for col, values in config["drop_downs"].items():
        options = values.get("options", [])
        prompt = values.get("prompt", "")
        title = values.get("title", "")
        add_drop_down_col(
            sheet=summary_sheet,
            header_name=col,
            dropdown_options=options,
            prompt=prompt,
            title=title,
        )
    format_summary_sheet(summary_sheet, data, col_widths)
