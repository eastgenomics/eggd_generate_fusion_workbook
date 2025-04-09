"""utilities for formating summary sheet
"""

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, DEFAULT_FONT, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .utils import generate_varsome_url

DEFAULT_FONT.name = "Calibri"
DEFAULT_FONT.size = 11


def add_extra_columns(
    worksheet: Worksheet, extra_cols: dict[str, str], start: int = 1
) -> None:
    """
    Inserts additional columns with formulas at the beginning of a sheet

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where columns will be added.
    extra_cols : dict[str, str]
        A mapping of column names to Excel formulas
    start : int, optional
        The column index (1-based) where extra columns should be inserted.
        Defaults to 1 (beginning of the sheet).

    Returns
    -------
    None
    """

    num_cols = len(extra_cols)
    worksheet.insert_cols(start, amount=num_cols)

    for i, (col, formula) in enumerate(extra_cols.items(), start=start):
        worksheet.cell(row=1, column=i, value=col)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=i, value=formula.replace("{row}", str(row)))


def apply_header_format(worksheet: Worksheet) -> None:
    """
    Applies bold formatting to all headers in the sheet.

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where header formatting will be applied.

    Returns
    -------
    None
    """
    header_font = Font(bold=True)
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font


def adjust_column_widths(
    worksheet: Worksheet, min_width: int = 14, max_width: int = 40
) -> None:
    """
    Adjusts all column widths in given sheet for better visibility.

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where column widths will be adjusted.
    min_width : int
        The min width a column should have.
    max_width : int
        The max width a column should have.

    Returns
    -------
    None
    """
    for col_cells in worksheet.columns:
        col_letter = col_cells[0].column_letter
        _max = min_width

        for cell in col_cells:
            val = cell.value
            if val and not (isinstance(val, str) and val.startswith("=")):
                _max = max(_max, len(str(val)))

        # Set column width with a cap of max_width
        worksheet.column_dimensions[col_letter].width = min(_max + 2, max_width)


def set_tab_color(worksheet: Worksheet, hex_color: str) -> None:
    """Sets worksheet tab colour

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet to apply tab colour
    hex_color : str
        Hex colour code for the sheet tab
    """
    worksheet.sheet_properties.tabColor = hex_color


def style_borders(worksheet: Worksheet, style: str = "thin") -> None:
    """Apply style to all border of given sheet

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where style will be applied
    style : str, optional
        desired border style to apply, by default "thin"
    """
    border = Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style),
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border


def highlight_specimen_borders(
    worksheet: Worksheet, df: pd.DataFrame, index_col: str
) -> None:
    """Adds thick borders between specimen groups.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet where formating will be applied
    df : pd.DataFrame
        Data frame containing original summary data
    index_col : str
        Column to use for grouping"
    """
    thick_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thick"),
    )

    # Find last row of each specimen group
    last_rows = df.groupby(index_col).tail(1).index

    # Apply border to last row of each group
    for row in last_rows:
        excel_row = row + 2
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=excel_row, column=col).border = thick_border


def alternate_specimen_colors(
    worksheet: Worksheet,
    df: pd.DataFrame,
    index_col: str,
    stop_col: str = "FFPM",
    start_row_idx: int = 2,
) -> None:
    """Apply alternate row colors between specimen groups in summary sheet.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet where formating will be applied
    df : pd.DataFrame
        Data frame containing original summary data
    index_col : str
        Column name to use for grouping
    stop_col : str
        Column name where formating should end.
    start_row_idx : int
        Excel row index where formating should start from.
        Defaults 2 (when header row in at index 1)
    """
    blue_fill = PatternFill(
        start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid"
    )
    colors = [blue_fill, green_fill]

    col_letter = get_col_letter(worksheet, stop_col)
    if not col_letter:
        raise ValueError(f"stop_col '{stop_col}' not found in worksheet.")
    stop_idx = openpyxl.utils.column_index_from_string(col_letter)

    last_rows = df.groupby(index_col).tail(1).index
    current_row = start_row_idx

    for idx, last_row in enumerate(last_rows):
        fill = colors[idx % len(colors)]
        last_row = last_row + start_row_idx + 1
        for row in range(current_row, last_row):
            for col in range(1, stop_idx):
                worksheet.cell(row=row, column=col).fill = fill

        current_row = last_row


def align_column_cells(
    worksheet: Worksheet, col_letter: str, direction: str = "left"
) -> None:
    """
    Align all cells in a specified column to a given horizontal alignment.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The target Excel worksheet
    col_letter : str
        Column letter to format (e.g., 'A', 'D')
    direction : str, optional
        Horizontal alignment direction. Valid values:
        - 'left' (default)
        - 'center'
        - 'right'
        - 'justify'
        - 'distributed'
    """
    valid_directions = ["left", "center", "right", "justify", "distributed"]
    if direction.lower() not in valid_directions:
        raise ValueError(f"Invalid direction. Choose from: {valid_directions}")

    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    alignment = Alignment(horizontal=direction.lower())

    for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.alignment = alignment


def set_column_width(worksheet: Worksheet, col_letter: str, width: float) -> None:
    """
    Set fixed width for a specified column.
    For columns needing specific adjustment.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The target Excel worksheet
    col_letter : str
        Column letter to format (e.g., 'B', 'AA')
    width : float
        Column width in character units (Excel's standard measurement)
    """
    worksheet.column_dimensions[col_letter].width = width


def colour_hyperlinks(worksheet: Worksheet) -> None:
    """
    Set text colour to blue if text contains hyperlink

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The target Excel worksheet
    """
    for cells in worksheet.rows:
        for cell in cells:
            if "HYPERLINK" in str(cell.value):
                cell.font = Font(color="00007f", name=DEFAULT_FONT.name)


def add_drop_down_col(
    sheet: Worksheet,
    header_name: str,
    dropdown_options: list[str],
    prompt: str,
    title: str,
    start: int = 1,
    position: int | None = None,
) -> None:
    """
    Creates a dropdown column at at the specified position.

    Parameters
    ----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        current worksheet
    header_name : str
        Name of the new dropdown column
    dropdown_options : List[str]
        List containing drop-down items
    prompt: str
            prompt message for drop-down
    title: str
            title message for drop-down
    start : int
        Row index where sheets starts, defaults to 1
    position : int|None
        Column index to add dropdown or at the end if not given
    """
    ws = sheet
    last_col = ws.max_column

    # Insert at specified position or at the end
    col_idx = position if position else last_col + 1
    ws.insert_cols(col_idx)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws[f"{col_letter}{start}"] = header_name

    # Create data validation for dropdown
    options = f'"{",".join(dropdown_options)}"'

    dv = DataValidation(
        type="list",
        formula1=options,
        showDropDown=False,
        allow_blank=True,
    )
    dv.prompt = prompt
    dv.promptTitle = title
    ws.add_data_validation(dv)

    # Applly to all rows
    for row in range(start + 1, ws.max_row + 1):
        dv.add(ws[f"{col_letter}{row}"])

    dv.showInputMessage = True
    dv.showErrorMessage = True

    # adjust column width to accommodate longest string in dropdown items
    col_width = max(len(x) for x in dropdown_options + [header_name])
    set_column_width(ws, col_letter, width=col_width + 2)


def add_hyperlink(cell: openpyxl.cell.cell.Cell, url: str, text: str = None) -> None:
    """
    Add Excel hyperlink formula to cell

    Parameters
    ----------
    cell : openpyxl.cell.cell.Cell
        The target cell.
    url : str
        The URL to link to.
    text : str
        Text to display in the cell. Defaults to the cell value.
    """

    value = text if text else cell.value

    if url:
        cell.value = f'=HYPERLINK("{url}", "{value}")'


def add_breakpoint_hyperlinks(
    writer: pd.ExcelWriter,
    breakpoint_columns: tuple[str, str] = ("leftbreakpoint", "rightbreakpoint"),
    header_row: int = 1,
) -> None:
    """
    Apply VarSome hyperlinks to breakpoint columns across all sheets in the workbook.

    Parameters
    ----------
    writer : pd.ExcelWriter
        The pandas openpyxl ExcelWriter
    breakpoint_columns : tuple[str, str]
        Column names to check for breakpoints.
    header_row : int
        Row number where headers are located (1-based).
    """

    for sheet_name, worksheet in writer.sheets.items():
        max_col = worksheet.max_column
        max_row = worksheet.max_row

        # Get header row values
        headers = [
            worksheet.cell(row=header_row, column=col).value
            for col in range(1, max_col + 1)
        ]
        headers = [
            col.strip().lower() if isinstance(col, str) else "" for col in headers
        ]

        for bp_col in breakpoint_columns:
            if bp_col not in headers:
                continue

            col_idx = headers.index(bp_col) + 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            for row in range(header_row + 1, max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                value = cell.value
                if value and isinstance(value, str):
                    try:
                        url = generate_varsome_url(value)
                        add_hyperlink(cell, url, value)
                    except Exception as e:
                        print(
                            f"Could not process {value} at {sheet_name}!{col_letter}{row}: {e}"
                        )


def get_col_letter(worksheet: Worksheet, col_name: str) -> str:
    """
    Getting the column letter with specific col name

    Parameters
    ----------
    worksheet: openpyxl.worksheet.worksheet.Worksheet
            current sheet
    col_name: str
            name of column to get col letter
    Return
    -------
    str
        column letter for specific column name
    """
    col_letter = None
    for column_cell in worksheet.iter_cols(1, worksheet.max_column):
        if column_cell[0].value.strip() == col_name:
            col_letter = column_cell[0].column_letter

    return col_letter


def drop_column(worksheet: Worksheet, col_name: str) -> bool:
    """
    Deletes a column from the worksheet given a column name

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet from which to delete the column.
    col_name : str
        The header name of the column to delete.

    Returns
    -------
    bool
        True if the column was found and deleted, False otherwise.
    """
    col_letter = get_col_letter(worksheet, col_name)
    if not col_letter:
        print(f"Column '{col_name}' not found. No deletion performed.")
        return False

    col_index = openpyxl.utils.column_index_from_string(col_letter)
    worksheet.delete_cols(col_index)
    print(f"Deleted column '{col_name}' (column {col_letter}).")
    return True


def rotate_headers(sheet: Worksheet, header_row: int = 1, degree: int = 90) -> None:
    """Rotates values in header row by given degree (90 -- vertically).

    Parameters
    ----------
    sheet : Worksheet
        The sheet where headers will be rotated.
    header_row : int, optional
        The row index containing headers (1-based), by default 1
    degree : int, optional
        The degree of rotation to apply (0-180), by default 90
    """
    alignment = Alignment(textRotation=degree, vertical="bottom", horizontal="center")
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.alignment = alignment
        cell.value = f" {cell.value}"

    sheet.row_dimensions[header_row].height = 130


def write_df_to_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    tab_color: str = "000000",
    extra_cols: dict[str, str] = None,
    include_index: bool = False,
) -> None:
    """Writes a Pandas DataFrame to an Excel sheet with formatting.

    Parameters
    ----------
    writer : pd.ExcelWriter
        The Excel writer object to write the sheet into
    df : pd.DataFrame
        The DataFrame containing the data.
    sheet_name : str
        Name of the Excel sheet
    tab_color : str, optional
        Hex colour code for the sheet tab. Defaults to black ("000000")
    extra_cols : dict[str, str], optional
        A mapping of new column names to excel formulas. Example:
            {
                "Specimen": "=MID(C{row},11,10)",
                "EPIC": "=VLOOKUP(A{row},EPIC!AJ:AK,2,0)"
            }
    include_index : bool, optional
        Wether to write index of df to sheet. Defaults to False
    """
    df.to_excel(writer, sheet_name=sheet_name, index=include_index)
    worksheet = writer.sheets[sheet_name]

    set_tab_color(worksheet, tab_color)

    # Add extra columns if provided
    if extra_cols:
        add_extra_columns(worksheet, extra_cols)

    adjust_column_widths(worksheet)


def format_workbook(writer: pd.ExcelWriter) -> None:
    """apply formatting to entire workbook

    Parameters
    ----------
    writer : pd.ExcelWriter
    The pandas openpyxl ExcelWriter
    """

    workbook = writer.book

    add_breakpoint_hyperlinks(writer)

    for sheet in workbook.worksheets:
        colour_hyperlinks(sheet)
        apply_header_format(sheet)

        # other general formating to follow here ...
