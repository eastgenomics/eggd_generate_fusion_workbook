"""general and I/O utilities
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

import dxpy
from dxpy import DXDataObject


def read_dxfile(
    dxfile: DXDataObject, sep: str = "\t", include_fname: bool = True
) -> pd.DataFrame:
    """reads a DNAnexus file object into a pandas dataframe

    Parameters
    ----------
    dxfile : DXDataObject
        An instance of DXDataObject; behaves like a python file object.
    sep : str
        Delimeter of data values in files. Defaults to  "\t"
    include_fname : bool
        Specifies whether to set file name as first column. Defaults to True

    Returns
    -------
    pd.DataFrame
        An instance of DataFrame with file content
    """

    df = pd.read_csv(dxfile, sep=sep)

    if include_fname:
        fname = dxfile.name
        df.insert(0, "file_name", fname)

    return df


def _add_extra_columns(
    worksheet: Worksheet, extra_cols: dict[str, str], start: int = 1
) -> None:
    """
    Inserts additional columns with formulas at the beginning of a sheet

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where columns will be added.
    extra_cols : dict[str, str]
        A mapping of column names to Excel formulas
    start : int, optional
        The column index (1-based) where extra columns should be inserted.
        Defaults to 1 (beginning of the sheet).

    Returns
    -------
    None
    """

    num_cols = len(extra_cols)
    worksheet.insert_cols(start, amount=num_cols)

    for i, (col, formula) in enumerate(extra_cols.items(), start=start):
        worksheet.cell(row=1, column=i, value=col)
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=i, value=formula.replace("{row}", str(row)))


def _apply_header_format(worksheet) -> None:
    """
    Applies bold formatting to all headers in the sheet.

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where header formatting will be applied.

    Returns
    -------
    None
    """
    header_font = Font(bold=True)
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font


def _adjust_column_widths(worksheet, min_width=14, max_width=40):
    """
    Adjusts column widths for better visibility.

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where column widths will be adjusted.
    min_width : int
        The min width a column should have.
    max_width : int
        The max width a column should have.

    Returns
    -------
    None
    """
    for col_cells in worksheet.columns:
        col_letter = col_cells[0].column_letter
        _max = min_width

        for cell in col_cells:
            val = cell.value
            if val and not (isinstance(val, str) and val.startswith("=")):
                _max = max(_max, len(str(val)))

        # Set column width with a cap of max_width
        worksheet.column_dimensions[col_letter].width = min(_max + 2, max_width)


def _set_tab_color(worksheet, hex_color: str) -> None:
    """Sets worksheet tab colour

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet to apply tab colour
    hex_color : str
        Hex colour code for the sheet tab
    """
    worksheet.sheet_properties.tabColor = hex_color


def highlight_max_ffpm(worksheet, source_df, ffpm_col="FFPM", index_col="SPECIMEN"):
    """
    Conditionally formats rows with maximum FFPM value per specimen(index).

    Parameters
    ----------
    worksheet : Worksheet
        The worksheet where conditional formatting will be applied
    source_df : pandas.DataFrame
        The source data written to sheet
    ffpm_col : str, optional
        Name of the FFPM column in the DataFrame, default is "FFPM"
    index_col : str, optional
        Name of the column to use for grouping ffpm, default is "SPECIMEN"

    Returns
    -------
    None
    """

    # Create a light green fill pattern for highlighting
    green_fill = PatternFill(
        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
    )

    # Reset index if it's a multi-index df to make processing easier
    if isinstance(source_df.index, pd.MultiIndex):
        df = source_df.reset_index()
    else:
        df = source_df.copy()

    # Find the max FFPM per specimen
    max_ffpm = df.groupby(index_col)[ffpm_col].transform("max")
    max_rows = df[(df[ffpm_col] == max_ffpm)].index.tolist()

    # Apply conditional formatting
    for idx in max_rows:
        row_idx = idx + 2  # adjust for 0-indexing in pandas
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = green_fill


def write_df_to_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    tab_color: str = "000000",
    extra_cols: dict[str, str] = None,
    include_index: bool = False,
) -> None:
    """Writes a Pandas DataFrame to an Excel sheet with formatting.

    Parameters
    ----------
    writer : pd.ExcelWriter
        The Excel writer object to write the sheet into
    df : pd.DataFrame
        The DataFrame containing the data.
    sheet_name : str
        Name of the Excel sheet
    tab_color : str, optional
        Hex colour code for the sheet tab. Defaults to black ("000000")
    extra_cols : dict[str, str], optional
        A mapping of new column names to excel formulas. Example:
            {
                "Specimen": "=MID(C{row},11,10)",
                "EPIC": "=VLOOKUP(A{row},EPIC!AJ:AK,2,0)"
            }
    include_index : bool, optional
        Wether to write index of df to sheet. Defaults to False
    """
    df.to_excel(writer, sheet_name=sheet_name, index=include_index)
    worksheet = writer.sheets[sheet_name]

    _set_tab_color(worksheet, tab_color)

    # Add extra columns if provided
    if extra_cols:
        _add_extra_columns(worksheet, extra_cols)

    _apply_header_format(worksheet)
    _adjust_column_widths(worksheet)


def create_pivot_table(
    df: pd.DataFrame,
    pivot_config: dict,
) -> pd.DataFrame:
    """Creates pivot table from a dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        Source dataframe to create pivot from
    pivot_config : dict
        Dictionary defining pivot structure:
            - index: Columns to group by (list for MultiIndex)
            - columns: Columns for column grouping
            - values: Value columns to aggregate
            - aggfunc: Aggregation function
            - add_total_row: Whether to add totals (default False)

    Returns
    ------
    pd.DataFrame
        created pivot table
    """

    # Create pivot table
    pivot_df = df.pivot_table(
        index=pivot_config["index"],
        columns=pivot_config["columns"],
        values=pivot_config["values"],
        aggfunc=pivot_config["aggfunc"],
        margins=pivot_config.get("add_total_row", False),
        margins_name="Total",
    )

    return pivot_df


def get_project_info() -> tuple[str, str]:
    """Get the project name for naming output file

    Returns
    -------
    tuple[str, str]
        Name and ID of DNAnexus project
    """

    project_id = os.environ.get("DX_PROJECT_CONTEXT_ID")

    # Get name of project for output naming
    project_name = dxpy.describe(project_id)["name"]
    project_name = "_".join(project_name.split("_")[1:])

    return project_name, project_id
